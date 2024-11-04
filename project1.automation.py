import openpyxl as xl   # Import the openpyxl module
import os
from openpyxl.chart import Reference, BarChart

def process_workbook(filename):

    wb = xl.load_workbook(filename)  # Load the workbook
    s1 = wb['Sheet1']  # Get the sheet object
    s1.cell(1,4,'New price')
    for row in range (2, s1.max_row + 1):
        cell = s1.cell(row, 3)
        s1.cell(row,4, cell.value * 0.9) 
    ref = Reference(s1,
                    min_row=2,
                    max_row=s1.max_row,
                    min_col=4,
                    max_col=4)
    chart = BarChart()
    chart.add_data(ref)
    s1.add_chart(chart, 'e2')
    wb.save(filename)


def process_folder(path):
    files = os.listdir(path)
 
    for file in files:
        process_workbook(path + '/' + file)




path1 = '/Users/mma2998/Library/CloudStorage/OneDrive-MassMutual/Documents/GitHub/desktop-tutorial/ex1'

process_folder(path1)








